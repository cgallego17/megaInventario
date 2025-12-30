from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db import transaction, models
from django.utils import timezone
from django import forms
import csv
import pandas as pd
from io import BytesIO

from .models import ComparativoInventario, InventarioSistema, ItemComparativo
from .forms import ComparativoInventarioForm, InventarioSistemaForm
from productos.models import Producto
from conteo.models import Conteo, ItemConteo


@login_required
def lista_comparativos(request):
    """Lista todos los comparativos - Redirige al único comparativo si existe"""
    comparativo = ComparativoInventario.objects.first()
    
    if comparativo:
        return redirect('comparativos:detalle', pk=comparativo.pk)
    else:
        # Si no existe, mostrar lista vacía con opción de crear
        return render(request, 'comparativos/lista.html', {'comparativos': []})


@login_required
def crear_comparativo(request):
    """Crea un nuevo comparativo - Solo permite uno, automáticamente suma todos los conteos"""
    # Verificar si ya existe un comparativo
    comparativo_existente = ComparativoInventario.objects.first()
    
    if comparativo_existente:
        messages.info(request, 'Ya existe un comparativo. Redirigiendo al existente.')
        return redirect('comparativos:detalle', pk=comparativo_existente.pk)
    
    if request.method == 'POST':
        form = ComparativoInventarioForm(request.POST)
        if form.is_valid():
            comparativo = form.save(commit=False)
            comparativo.usuario = request.user
            comparativo.save()
            
            # Automáticamente procesar con todos los conteos finalizados
            from django.db.models import Sum
            conteos_finalizados = Conteo.objects.filter(estado='finalizado')
            
            if conteos_finalizados.exists():
                # Agrupar items por producto y sumar cantidades
                items_agregados = ItemConteo.objects.filter(
                    conteo__in=conteos_finalizados
                ).values('producto').annotate(
                    cantidad_total=Sum('cantidad')
                )
                
                cantidad_por_producto = {item['producto']: item['cantidad_total'] for item in items_agregados}
                
                # Obtener todos los productos (activos e inactivos)
                productos = Producto.objects.all()
                
                with transaction.atomic():
                    for producto in productos:
                        item, created = ItemComparativo.objects.get_or_create(
                            comparativo=comparativo,
                            producto=producto
                        )
                        item.cantidad_fisico = cantidad_por_producto.get(producto.id, 0)
                        item.calcular_diferencias()
                
                conteos_usados = conteos_finalizados.count()
                messages.success(request, f'Comparativo creado y procesado. Se sumaron {conteos_usados} conteo(s) finalizado(s).')
            else:
                messages.warning(request, 'Comparativo creado, pero no hay conteos finalizados para procesar.')
            
            return redirect('comparativos:detalle', pk=comparativo.pk)
    else:
        form = ComparativoInventarioForm()
    
    return render(request, 'comparativos/form.html', {'form': form, 'titulo': 'Crear Comparativo'})


@login_required
def subir_inventario(request, pk):
    """Sube inventarios de los sistemas (opcional)"""
    comparativo = get_object_or_404(ComparativoInventario.objects.select_related('conteo', 'usuario'), pk=pk)
    inventarios = comparativo.inventarios.all()
    
    sistema1_subido = inventarios.filter(sistema='sistema1').first()
    sistema2_subido = inventarios.filter(sistema='sistema2').first()
    
    # Estadísticas de items cargados
    items_sistema1 = comparativo.items.exclude(cantidad_sistema1=0).count() if sistema1_subido else 0
    items_sistema2 = comparativo.items.exclude(cantidad_sistema2=0).count() if sistema2_subido else 0
    
    # Procesar automáticamente con todos los conteos si no hay items o faltan productos
    productos = Producto.objects.all()
    productos_en_comparativo = set(comparativo.items.values_list('producto_id', flat=True))
    productos_faltantes = [p for p in productos if p.id not in productos_en_comparativo]
    
    if comparativo.items.count() == 0 or productos_faltantes:
        from django.db.models import Sum
        conteos_finalizados = Conteo.objects.filter(estado='finalizado')
        
        if conteos_finalizados.exists():
            items_agregados = ItemConteo.objects.filter(
                conteo__in=conteos_finalizados
            ).values('producto').annotate(
                cantidad_total=Sum('cantidad')
            )
            
            cantidad_por_producto = {item['producto']: item['cantidad_total'] for item in items_agregados}
            
            # Crear o actualizar items para todos los productos
            with transaction.atomic():
                for producto in productos:
                    item, created = ItemComparativo.objects.get_or_create(
                        comparativo=comparativo,
                        producto=producto
                    )
                    item.cantidad_fisico = cantidad_por_producto.get(producto.id, 0)
                    item.calcular_diferencias()
        else:
            # Si no hay conteos finalizados, crear items con cantidad 0 para todos los productos
            with transaction.atomic():
                for producto in productos:
                    item, created = ItemComparativo.objects.get_or_create(
                        comparativo=comparativo,
                        producto=producto
                    )
                    if created:
                        item.cantidad_fisico = 0
                        item.calcular_diferencias()
    
    if request.method == 'POST':
        sistema = request.POST.get('sistema')
        
        # Validar que se haya seleccionado un sistema
        if not sistema or sistema not in ['sistema1', 'sistema2']:
            messages.error(request, 'Debe seleccionar un sistema válido.')
            form = InventarioSistemaForm()
        else:
            form = InventarioSistemaForm(request.POST, request.FILES)
            
            if form.is_valid():
                try:
                    # Validar que se haya subido un archivo
                    if 'archivo' not in request.FILES:
                        messages.error(request, 'Debe seleccionar un archivo para subir.')
                    else:
                        archivo = request.FILES['archivo']
                        
                        # Validar extensión del archivo
                        nombre_archivo = archivo.name.lower()
                        if not (nombre_archivo.endswith('.xlsx') or nombre_archivo.endswith('.xls') or nombre_archivo.endswith('.csv')):
                            messages.error(request, 'El archivo debe ser Excel (.xlsx, .xls) o CSV (.csv).')
                        else:
                            # Pasar el sistema seleccionado para que el form pueda detectar la columna correcta
                            inventario_data, nombre_sistema_archivo = form.procesar_archivo(archivo, sistema=sistema)
                            
                            # Validar que se hayan procesado productos
                            if not inventario_data:
                                messages.warning(request, 'El archivo no contiene productos válidos para procesar.')
                            else:
                                # Actualizar el nombre del sistema si se encontró en el archivo
                                if nombre_sistema_archivo:
                                    if sistema == 'sistema1':
                                        comparativo.nombre_sistema1 = nombre_sistema_archivo
                                    elif sistema == 'sistema2':
                                        comparativo.nombre_sistema2 = nombre_sistema_archivo
                                    comparativo.save()
                                    messages.info(request, f'Nombre del sistema actualizado a: {nombre_sistema_archivo}')
                                
                                # Guardar o actualizar inventario del sistema
                                inventario_sistema, created = InventarioSistema.objects.update_or_create(
                                    comparativo=comparativo,
                                    sistema=sistema,
                                    defaults={'archivo': archivo}
                                )
                                
                                # Procesar y crear items comparativos
                                productos_encontrados = 0
                                productos_no_encontrados = 0
                                
                                with transaction.atomic():
                                    for codigo_barras, cantidad in inventario_data.items():
                                        try:
                                            # Buscar producto por código de barras o código interno (todos los productos)
                                            producto = Producto.objects.filter(
                                                models.Q(codigo_barras=codigo_barras) | models.Q(codigo=codigo_barras)
                                            ).first()
                                            
                                            if producto:
                                                item, item_created = ItemComparativo.objects.get_or_create(
                                                    comparativo=comparativo,
                                                    producto=producto
                                                )
                                                
                                                if sistema == 'sistema1':
                                                    item.cantidad_sistema1 = cantidad
                                                elif sistema == 'sistema2':
                                                    item.cantidad_sistema2 = cantidad
                                                
                                                item.save()
                                                productos_encontrados += 1
                                            else:
                                                productos_no_encontrados += 1
                                        except Exception as e:
                                            # Log del error pero continuar con otros productos
                                            productos_no_encontrados += 1
                                            continue
                                
                                mensaje = f'Inventario del {form.cleaned_data.get("sistema", sistema)} cargado exitosamente. '
                                mensaje += f'{productos_encontrados} productos procesados.'
                                if productos_no_encontrados > 0:
                                    mensaje += f' {productos_no_encontrados} productos no encontrados en el sistema.'
                                messages.success(request, mensaje)
                                return redirect('comparativos:subir_inventario', pk=comparativo.pk)
                
                except forms.ValidationError as e:
                    # Errores de validación del formulario
                    messages.error(request, f'Error de validación: {str(e)}')
                except Exception as e:
                    # Otros errores
                    import traceback
                    error_detail = str(e)
                    messages.error(request, f'Error al procesar el archivo: {error_detail}')
            else:
                # Formulario no válido
                if form.errors:
                    for field, errors in form.errors.items():
                        for error in errors:
                            messages.error(request, f'{field}: {error}')
    else:
        form = InventarioSistemaForm()
    
    return render(request, 'comparativos/subir_inventario.html', {
        'comparativo': comparativo,
        'form': form,
        'sistema1_subido': sistema1_subido,
        'sistema2_subido': sistema2_subido,
        'items_sistema1': items_sistema1,
        'items_sistema2': items_sistema2,
    })


@login_required
def procesar_comparativo(request, pk):
    """Procesa el comparativo cargando datos del conteo físico - Prioriza reconteos relacionados con este comparativo"""
    comparativo = get_object_or_404(ComparativoInventario, pk=pk)
    
    from django.db.models import Sum
    
    # Buscar conteos de reconteo relacionados con este comparativo
    conteos_reconteo = Conteo.objects.filter(
        estado='finalizado',
        observaciones__contains=f'Conteo creado desde comparativo "{comparativo.nombre}"'
    )
    
    # Obtener TODOS los conteos finalizados para productos que no están en reconteos
    conteos_finalizados = Conteo.objects.filter(estado='finalizado')
    
    if not conteos_finalizados.exists():
        messages.warning(request, 'No hay conteos finalizados. El comparativo se procesará con cantidad física 0.')
    
    # Obtener productos de reconteos (si existen)
    productos_reconteo_ids = set()
    if conteos_reconteo.exists():
        for conteo_reconteo in conteos_reconteo:
            if conteo_reconteo.observaciones and 'Productos:' in conteo_reconteo.observaciones:
                try:
                    productos_str = conteo_reconteo.observaciones.split('Productos:')[1].strip()
                    productos_ids = [int(pid.strip()) for pid in productos_str.split(',') if pid.strip().isdigit()]
                    productos_reconteo_ids.update(productos_ids)
                except (ValueError, AttributeError):
                    pass
    
    # Agrupar items por producto: reconteos tienen prioridad
    cantidad_por_producto = {}
    
    if productos_reconteo_ids:
        # Items de reconteos (solo conteos de reconteo relacionados)
        items_reconteo = ItemConteo.objects.filter(
            conteo__in=conteos_reconteo,
            producto_id__in=productos_reconteo_ids
        ).values('producto').annotate(
            cantidad_total=Sum('cantidad')
        )
        for item in items_reconteo:
            cantidad_por_producto[item['producto']] = item['cantidad_total']
        
        # Items de productos normales (excluyendo productos de reconteo)
        items_normales = ItemConteo.objects.filter(
            conteo__in=conteos_finalizados
        ).exclude(
            producto_id__in=productos_reconteo_ids
        ).values('producto').annotate(
            cantidad_total=Sum('cantidad')
        )
    else:
        # Si no hay reconteos, usar todos los conteos finalizados
        items_normales = ItemConteo.objects.filter(
            conteo__in=conteos_finalizados
        ).values('producto').annotate(
            cantidad_total=Sum('cantidad')
        )
    
    # Agregar cantidades de productos normales
    for item in items_normales:
        producto_id = item['producto']
        if producto_id not in cantidad_por_producto:
            cantidad_por_producto[producto_id] = item['cantidad_total']
    
    # Obtener todos los productos (activos e inactivos) para asegurar que todos estén en el comparativo
    productos = Producto.objects.all()
    
    with transaction.atomic():
        for producto in productos:
            item, created = ItemComparativo.objects.get_or_create(
                comparativo=comparativo,
                producto=producto
            )
            # Asignar la cantidad total sumada de todos los conteos finalizados
            item.cantidad_fisico = cantidad_por_producto.get(producto.id, 0)
            item.calcular_diferencias()
    
    # Mensaje informativo
    conteos_usados = conteos_finalizados.count()
    if conteos_reconteo.exists():
        mensaje = f'Comparativo procesado exitosamente. Se usaron {conteos_reconteo.count()} reconteo(s) para productos específicos y {conteos_usados} conteo(s) finalizado(s) en total.'
    else:
        mensaje = f'Comparativo procesado exitosamente. Se sumaron {conteos_usados} conteo(s) finalizado(s).'
    
    messages.success(request, mensaje)
    return redirect('comparativos:detalle', pk=comparativo.pk)


@login_required
def detalle_comparativo(request, pk):
    """Muestra el detalle del comparativo - Asegura que todos los productos estén incluidos"""
    comparativo = get_object_or_404(ComparativoInventario.objects.select_related('conteo', 'usuario').prefetch_related('conteo__parejas'), pk=pk)
    
    # Asegurar que todos los productos (activos e inactivos) tengan un item en el comparativo
    # Optimizado: usar bulk_create en lugar de crear uno por uno
    productos = Producto.objects.all().only('id')
    productos_en_comparativo = set(comparativo.items.values_list('producto_id', flat=True))
    
    # Crear items para productos que no están en el comparativo usando bulk_create
    productos_faltantes = [p for p in productos if p.id not in productos_en_comparativo]
    if productos_faltantes:
        items_nuevos = [
            ItemComparativo(
                comparativo=comparativo,
                producto=producto,
                cantidad_sistema1=0,
                cantidad_sistema2=0,
                cantidad_fisico=0,
                diferencia_sistema1=0,
                diferencia_sistema2=0
            )
            for producto in productos_faltantes
        ]
        ItemComparativo.objects.bulk_create(items_nuevos, ignore_conflicts=True)
    
    # Obtener todos los items (incluyendo los recién creados) con select_related para optimizar
    # Ordenar por marca primero, luego por nombre
    items = comparativo.items.all().select_related('producto').order_by('producto__marca', 'producto__nombre')
    
    # Obtener información sobre los conteos finalizados que se están usando
    # Optimizado: usar agregaciones de base de datos
    from django.db.models import Sum, Count
    conteos_finalizados = Conteo.objects.filter(estado='finalizado').order_by('numero_conteo', '-fecha_fin')
    conteos_info = []
    for conteo in conteos_finalizados:
        # Usar agregación en lugar de sum() en Python
        total_cantidad_conteo = conteo.items.aggregate(total=Sum('cantidad'))['total'] or 0
        total_items_conteo = conteo.items.count()
        conteos_info.append({
            'conteo': conteo,
            'total_items': total_items_conteo,
            'total_cantidad': total_cantidad_conteo,
        })
    
    # Estadísticas - optimizado usando agregaciones de base de datos
    total_items = items.count()
    items_con_diferencia_s1 = items.exclude(diferencia_sistema1=0).count()
    items_con_diferencia_s2 = items.exclude(diferencia_sistema2=0).count()
    
    # Usar agregaciones de base de datos para sumas en lugar de sum() en Python
    from django.db.models import F, Sum as SumAgg
    estadisticas = items.aggregate(
        total_cantidad_sistema1=SumAgg('cantidad_sistema1'),
        total_cantidad_sistema2=SumAgg('cantidad_sistema2'),
        total_cantidad_fisico=SumAgg('cantidad_fisico'),
        total_valor_sistema1=SumAgg(F('producto__precio') * F('cantidad_sistema1')),
        total_valor_sistema2=SumAgg(F('producto__precio') * F('cantidad_sistema2')),
        total_valor_fisico=SumAgg(F('producto__precio') * F('cantidad_fisico')),
    )
    
    total_cantidad_sistema1 = estadisticas['total_cantidad_sistema1'] or 0
    total_cantidad_sistema2 = estadisticas['total_cantidad_sistema2'] or 0
    total_cantidad_fisico = estadisticas['total_cantidad_fisico'] or 0
    total_valor_sistema1 = float(estadisticas['total_valor_sistema1'] or 0)
    total_valor_sistema2 = float(estadisticas['total_valor_sistema2'] or 0)
    total_valor_fisico = float(estadisticas['total_valor_fisico'] or 0)
    
    diferencia_valor_s1 = total_valor_fisico - total_valor_sistema1
    diferencia_valor_s2 = total_valor_fisico - total_valor_sistema2
    diferencia_cantidad_s1 = total_cantidad_fisico - total_cantidad_sistema1
    diferencia_cantidad_s2 = total_cantidad_fisico - total_cantidad_sistema2
    
    # Obtener parejas activas para el selector
    from usuarios.models import ParejaConteo
    parejas_activas = ParejaConteo.objects.filter(activa=True).order_by('usuario_1__username', 'usuario_2__username')
    
    # Obtener conteos de reconteo existentes (creados desde comparativos, sin parejas asignadas)
    from django.db.models import Count
    conteos_recontar_existentes = Conteo.objects.filter(
        estado='en_proceso',
        observaciones__contains='Conteo creado desde comparativo'
    ).annotate(
        num_parejas=Count('parejas')
    ).filter(
        num_parejas=0
    ).distinct().order_by('-fecha_creacion')
    
    return render(request, 'comparativos/detalle.html', {
        'comparativo': comparativo,
        'items': items,
        'total_items': total_items,
        'items_con_diferencia_s1': items_con_diferencia_s1,
        'items_con_diferencia_s2': items_con_diferencia_s2,
        'total_cantidad_sistema1': total_cantidad_sistema1,
        'total_cantidad_sistema2': total_cantidad_sistema2,
        'total_cantidad_fisico': total_cantidad_fisico,
        'total_valor_sistema1': total_valor_sistema1,
        'total_valor_sistema2': total_valor_sistema2,
        'total_valor_fisico': total_valor_fisico,
        'diferencia_cantidad_s1': diferencia_cantidad_s1,
        'diferencia_cantidad_s2': diferencia_cantidad_s2,
        'diferencia_valor_s1': diferencia_valor_s1,
        'diferencia_valor_s2': diferencia_valor_s2,
        'conteos_finalizados': conteos_finalizados,
        'conteos_info': conteos_info,
        'parejas_activas': parejas_activas,
        'conteos_recontar_existentes': conteos_recontar_existentes,
    })


@login_required
def exportar_comparativo(request, pk):
    """Exporta el comparativo a Excel"""
    comparativo = get_object_or_404(ComparativoInventario, pk=pk)
    # Ordenar por marca primero, luego por nombre
    items = comparativo.items.all().select_related('producto').order_by('producto__marca', 'producto__nombre')
    
    # Preparar datos para el DataFrame
    datos = []
    for item in items:
        datos.append({
            'Código de Barras': item.producto.codigo_barras,
            'Código': item.producto.codigo or '',
            'Marca': item.producto.marca or '',
            'Producto': item.producto.nombre,
            'Atributo': item.producto.atributo or '',
            'Precio Unitario': float(item.get_precio()),
            f'Cantidad {comparativo.nombre_sistema1 or "Sistema 1"}': item.cantidad_sistema1,
            f'Valor {comparativo.nombre_sistema1 or "Sistema 1"}': float(item.get_valor_sistema1()),
            f'Cantidad {comparativo.nombre_sistema2 or "Sistema 2"}': item.cantidad_sistema2,
            f'Valor {comparativo.nombre_sistema2 or "Sistema 2"}': float(item.get_valor_sistema2()),
            'Cantidad Físico': item.cantidad_fisico,
            'Valor Físico': float(item.get_valor_fisico()),
            f'Diferencia Cantidad {comparativo.nombre_sistema1 or "Sistema 1"}': item.diferencia_sistema1,
            f'Diferencia Valor {comparativo.nombre_sistema1 or "Sistema 1"}': float(item.get_diferencia_valor_sistema1()),
            f'Diferencia Cantidad {comparativo.nombre_sistema2 or "Sistema 2"}': item.diferencia_sistema2,
            f'Diferencia Valor {comparativo.nombre_sistema2 or "Sistema 2"}': float(item.get_diferencia_valor_sistema2())
        })
    
    # Crear DataFrame
    df = pd.DataFrame(datos)
    
    # Agregar fila de totales
    if len(datos) > 0:
        totales = {
            'Código de Barras': '',
            'Código': '',
            'Marca': '',
            'Producto': 'TOTALES',
            'Atributo': '',
            'Precio Unitario': '',
            f'Cantidad {comparativo.nombre_sistema1 or "Sistema 1"}': sum(item.cantidad_sistema1 for item in items),
            f'Valor {comparativo.nombre_sistema1 or "Sistema 1"}': sum(item.get_valor_sistema1() for item in items),
            f'Cantidad {comparativo.nombre_sistema2 or "Sistema 2"}': sum(item.cantidad_sistema2 for item in items),
            f'Valor {comparativo.nombre_sistema2 or "Sistema 2"}': sum(item.get_valor_sistema2() for item in items),
            'Cantidad Físico': sum(item.cantidad_fisico for item in items),
            'Valor Físico': sum(item.get_valor_fisico() for item in items),
            f'Diferencia Cantidad {comparativo.nombre_sistema1 or "Sistema 1"}': sum(item.diferencia_sistema1 for item in items),
            f'Diferencia Valor {comparativo.nombre_sistema1 or "Sistema 1"}': sum(item.get_diferencia_valor_sistema1() for item in items),
            f'Diferencia Cantidad {comparativo.nombre_sistema2 or "Sistema 2"}': sum(item.diferencia_sistema2 for item in items),
            f'Diferencia Valor {comparativo.nombre_sistema2 or "Sistema 2"}': sum(item.get_diferencia_valor_sistema2() for item in items)
        }
        df_totales = pd.DataFrame([totales])
        df = pd.concat([df, df_totales], ignore_index=True)
    
    # Crear archivo Excel en memoria
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Comparativo', index=False)
        
        # Obtener la hoja para formatear
        worksheet = writer.sheets['Comparativo']
        
        # Importar estilos de openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 20,  # Código de Barras
            'B': 15,  # Código
            'C': 15,  # Marca
            'D': 40,  # Producto
            'E': 15,  # Atributo
            'F': 15,  # Precio Unitario
            'G': 18,  # Cantidad Sistema 1
            'H': 18,  # Valor Sistema 1
            'I': 18,  # Cantidad Sistema 2
            'J': 18,  # Valor Sistema 2
            'K': 18,  # Cantidad Físico
            'L': 18,  # Valor Físico
            'M': 22,  # Diferencia Cantidad Sistema 1
            'N': 22,  # Diferencia Valor Sistema 1
            'O': 22,  # Diferencia Cantidad Sistema 2
            'P': 22,  # Diferencia Valor Sistema 2
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Formatear encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Formatear filas de datos
        for row_idx in range(2, len(df) + 1):
            for col_idx, col_letter in enumerate(column_widths.keys(), start=1):
                cell = worksheet[f'{col_letter}{row_idx}']
                cell.border = border
                
                # Formatear según el tipo de columna
                if col_letter in ['F', 'H', 'J', 'L', 'N', 'P']:  # Columnas de valores monetarios
                    if cell.value and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_letter in ['G', 'I', 'K', 'M', 'O']:  # Columnas de cantidades
                    if cell.value and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                else:  # Columnas de texto
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Formatear fila de totales (última fila)
        if len(df) > 0:
            total_row = len(df) + 1
            total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            total_font = Font(bold=True, size=11)
            
            for col_letter in column_widths.keys():
                cell = worksheet[f'{col_letter}{total_row}']
                cell.fill = total_fill
                cell.font = total_font
                cell.border = border
                
                if col_letter == 'D':  # Columna de Producto
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_letter in ['F', 'H', 'J', 'L', 'N', 'P']:  # Valores monetarios
                    if cell.value and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_letter in ['G', 'I', 'K', 'M', 'O']:  # Cantidades
                    if cell.value and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Agregar copyright al final
        copyright_row = len(df) + 2
        from datetime import datetime
        current_year = datetime.now().year
        worksheet.merge_cells(f'A{copyright_row}:P{copyright_row}')
        copyright_cell = worksheet[f'A{copyright_row}']
        copyright_cell.value = f'© {current_year} Todos los derechos reservados por megadominio.co'
        copyright_cell.alignment = Alignment(horizontal="center", vertical="center")
        copyright_font = Font(size=9, italic=True, color="808080")
        copyright_cell.font = copyright_font
    
    output.seek(0)
    
    # Crear la respuesta HTTP
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="comparativo_{comparativo.id}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    return response


@login_required
def descargar_ejemplo(request):
    """Descarga un archivo Excel de ejemplo para importación con TODOS los productos del sistema"""
    # Obtener TODOS los productos del sistema (activos e inactivos)
    productos = Producto.objects.all().order_by('nombre')
    
    # Crear datos con todos los productos - incluir columnas para ambos sistemas
    datos_ejemplo = []
    for producto in productos:
        datos_ejemplo.append({
            'codigo_barras': producto.codigo_barras,
            'codigo': producto.codigo or '',
            'nombre': producto.nombre,
            'marca': producto.marca or '',
            'atributo': producto.atributo or '',
            'cantidad': '',  # Vacío para que lo llenen
        })
    
    # Si no hay productos, usar datos de ejemplo genéricos
    if not datos_ejemplo:
        datos_ejemplo = [
            {'codigo_barras': '1234567890123', 'codigo': '', 'nombre': 'Producto Ejemplo 1', 'marca': 'Marca A', 'atributo': 'Ejemplo', 'cantidad': ''},
            {'codigo_barras': '9876543210987', 'codigo': '', 'nombre': 'Producto Ejemplo 2', 'marca': 'Marca B', 'atributo': 'Ejemplo', 'cantidad': ''},
            {'codigo_barras': '5555555555555', 'codigo': '', 'nombre': 'Producto Ejemplo 3', 'marca': 'Marca C', 'atributo': 'Ejemplo', 'cantidad': ''},
        ]
    
    df = pd.DataFrame(datos_ejemplo)
    
    # Detectar productos duplicados por codigo_barras o codigo
    # Crear una columna temporal para identificar duplicados
    df['_es_duplicado_cb'] = df.duplicated(subset=['codigo_barras'], keep=False)
    df['_es_duplicado_cod'] = df.duplicated(subset=['codigo'], keep=False) & (df['codigo'] != '')
    df['_es_duplicado'] = df['_es_duplicado_cb'] | df['_es_duplicado_cod']
    
    # Crear el archivo Excel en memoria
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # ===== HOJA 1: CONFIGURACIÓN =====
        config_data = {
            'Parámetro': ['Nombre del Sistema', 'Fecha de Inventario', 'Notas'],
            'Valor': ['Sistema 1', '', ''],
            'Descripción': [
                'Ingrese el nombre de este sistema (ej: SAP, Oracle, Sistema Legacy, etc.)',
                'Fecha del inventario (opcional)',
                'Notas adicionales (opcional)'
            ]
        }
        df_config = pd.DataFrame(config_data)
        df_config.to_excel(writer, sheet_name='Configuración', index=False)
        
        config_worksheet = writer.sheets['Configuración']
        config_worksheet.column_dimensions['A'].width = 25
        config_worksheet.column_dimensions['B'].width = 30
        config_worksheet.column_dimensions['C'].width = 60
        
        # Formatear hoja de configuración
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        config_header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        config_header_font = Font(bold=True, color="FFFFFF", size=12)
        config_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formatear encabezado de configuración
        for cell in config_worksheet[1]:
            cell.fill = config_header_fill
            cell.font = config_header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = config_border
        
        # Formatear filas de configuración
        for row_idx in range(2, len(df_config) + 2):
            for col_letter in ['A', 'B', 'C']:
                cell = config_worksheet[f'{col_letter}{row_idx}']
                cell.border = config_border
                if col_letter == 'A':
                    cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif col_letter == 'B':
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:  # Descripción
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                    cell.font = Font(italic=True, size=9, color="666666")
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Agregar instrucciones en la hoja de configuración
        instrucciones_row = len(df_config) + 3
        config_worksheet.merge_cells(f'A{instrucciones_row}:C{instrucciones_row + 4}')
        instrucciones_cell = config_worksheet[f'A{instrucciones_row}']
        instrucciones_text = (
            '📋 INSTRUCCIONES DE CONFIGURACIÓN:\n\n'
            '1. Complete el campo "Nombre del Sistema" con el nombre de su sistema (ej: SAP, Oracle, Sistema Legacy, etc.)\n'
            '2. Este nombre se usará automáticamente en el comparativo\n'
            '3. Si no completa el nombre, se usará "Sistema 1" o "Sistema 2" por defecto\n'
            '4. La fecha y notas son opcionales\n\n'
            '⚠️ IMPORTANTE: NO modifique los nombres de las columnas en esta hoja.'
        )
        instrucciones_cell.value = instrucciones_text
        instrucciones_cell.font = Font(size=10, color="333333")
        instrucciones_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        config_worksheet.row_dimensions[instrucciones_row].height = 120
        
        # ===== HOJA 2: INVENTARIO =====
        df_export = df.drop(columns=['_es_duplicado_cb', '_es_duplicado_cod', '_es_duplicado'])
        df_export.to_excel(writer, sheet_name='Inventario', index=False)
        
        worksheet = writer.sheets['Inventario']
        
        # Ajustar ancho de columnas
        worksheet.column_dimensions['A'].width = 20  # codigo_barras
        worksheet.column_dimensions['B'].width = 15  # codigo
        worksheet.column_dimensions['C'].width = 40  # nombre
        worksheet.column_dimensions['D'].width = 15  # marca
        worksheet.column_dimensions['E'].width = 15  # atributo
        worksheet.column_dimensions['F'].width = 18  # cantidad
        
        # Formatear encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Formatear filas y detectar duplicados
        from openpyxl.styles import Alignment as CellAlignment
        duplicate_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        duplicate_font = Font(color="CC0000", bold=True)
        
        for idx, row in df.iterrows():
            excel_row = idx + 2
            
            if row['_es_duplicado']:
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    cell = worksheet[f'{col}{excel_row}']
                    cell.fill = duplicate_fill
                    cell.font = duplicate_font
                    cell.border = border
                    if col == 'F':  # Columna de cantidad
                        cell.number_format = '0'
                        cell.alignment = CellAlignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = CellAlignment(horizontal="left", vertical="center")
            else:
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    cell = worksheet[f'{col}{excel_row}']
                    cell.border = border
                    if col == 'F':  # Columna de cantidad
                        cell.number_format = '0'
                        cell.alignment = CellAlignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = CellAlignment(horizontal="left", vertical="center")
        
        # Agregar instrucciones en la hoja de inventario
        total_duplicados = df['_es_duplicado'].sum()
        nota_row = len(datos_ejemplo) + 3
        worksheet.merge_cells(f'A{nota_row}:F{nota_row + 5}')
        nota_cell = worksheet[f'A{nota_row}']
        
        instrucciones = (
            '📋 INSTRUCCIONES DE USO:\n\n'
            '1. Complete la columna "cantidad" con los valores de inventario de cada producto\n'
            '2. Las columnas codigo_barras, codigo, nombre, marca y atributo son informativas y NO deben modificarse\n'
            '3. Puede usar este mismo archivo para ambos sistemas o crear archivos separados\n'
            '4. Si crea archivos separados, complete el nombre del sistema en la hoja "Configuración" de cada archivo\n\n'
        )
        
        if total_duplicados > 0:
            instrucciones += f'⚠️ ADVERTENCIA: Se detectaron {total_duplicados} productos con códigos duplicados (mismo código de barras o código interno). Estas filas están resaltadas en ROJO. Revise y corrija los duplicados antes de subir el archivo.'
        
        nota_cell.value = instrucciones
        nota_cell.font = Font(italic=True, size=9, color="666666")
        nota_cell.alignment = CellAlignment(horizontal="left", vertical="center", wrap_text=True)
        worksheet.row_dimensions[nota_row].height = 100
        
        # Agregar copyright en la hoja de Inventario
        copyright_row = len(df_export) + 2
        from datetime import datetime
        current_year = datetime.now().year
        worksheet.merge_cells(f'A{copyright_row}:F{copyright_row}')
        copyright_cell = worksheet[f'A{copyright_row}']
        copyright_cell.value = f'© {current_year} Todos los derechos reservados por megadominio.co'
        copyright_cell.alignment = Alignment(horizontal="center", vertical="center")
        copyright_font = Font(size=9, italic=True, color="808080")
        copyright_cell.font = copyright_font
        
        # Agregar copyright en la hoja de Configuración
        config_copyright_row = len(df_config) + 8
        config_worksheet.merge_cells(f'A{config_copyright_row}:C{config_copyright_row}')
        config_copyright_cell = config_worksheet[f'A{config_copyright_row}']
        config_copyright_cell.value = f'© {current_year} Todos los derechos reservados por megadominio.co'
        config_copyright_cell.alignment = Alignment(horizontal="center", vertical="center")
        config_copyright_cell.font = copyright_font
    
    output.seek(0)
    
    # Crear la respuesta HTTP
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="plantilla_inventario_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    return response


@login_required
def asignar_productos_recontar(request, pk):
    """Crea un nuevo conteo o agrega productos a un conteo existente para recontar"""
    comparativo = get_object_or_404(ComparativoInventario, pk=pk)
    
    if request.method == 'POST':
        producto_ids = request.POST.getlist('productos[]')
        accion = request.POST.get('accion', 'crear')  # 'crear' o 'agregar'
        conteo_id = request.POST.get('conteo_id')
        
        if not producto_ids:
            return JsonResponse({'success': False, 'error': 'Debe seleccionar al menos un producto.'})
        
        try:
            productos = Producto.objects.filter(id__in=producto_ids)
            
            if accion == 'agregar':
                # Agregar productos a un conteo existente
                if not conteo_id:
                    return JsonResponse({'success': False, 'error': 'Debe seleccionar un conteo existente.'})
                
                try:
                    conteo = Conteo.objects.get(pk=conteo_id, estado='en_proceso')
                    
                    # Verificar que el conteo fue creado desde un comparativo
                    if 'Conteo creado desde comparativo' not in (conteo.observaciones or ''):
                        return JsonResponse({'success': False, 'error': 'El conteo seleccionado no es válido para agregar productos.'})
                    
                    # Obtener productos actuales del conteo
                    productos_actuales = []
                    if conteo.observaciones and 'Productos:' in conteo.observaciones:
                        productos_str = conteo.observaciones.split('Productos:')[1].strip()
                        productos_actuales = [int(pid.strip()) for pid in productos_str.split(',') if pid.strip().isdigit()]
                    
                    # Agregar nuevos productos (sin duplicados)
                    productos_nuevos_ids = [int(pid) for pid in producto_ids]
                    productos_todos_ids = list(set(productos_actuales + productos_nuevos_ids))
                    
                    # Actualizar observaciones con todos los productos
                    productos_ids_str = ','.join(str(pid) for pid in productos_todos_ids)
                    conteo.observaciones = f'Conteo creado desde comparativo "{comparativo.nombre}" para recontar productos con diferencias. Productos: {productos_ids_str}'
                    conteo.usuario_modificador = request.user
                    conteo.save()
                    
                    productos_agregados = len([p for p in productos_nuevos_ids if p not in productos_actuales])
                    
                    return JsonResponse({
                        'success': True,
                        'message': f'{productos_agregados} producto(s) agregado(s) al conteo "{conteo.nombre}". Total de productos en el conteo: {len(productos_todos_ids)}.',
                        'conteo_id': conteo.pk
                    })
                except Conteo.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'Conteo no encontrado o no está en proceso.'})
            
            else:
                # Crear nuevo conteo
                nombre_conteo = request.POST.get('nombre_conteo', '').strip()
                numero_conteo = request.POST.get('numero_conteo')
                
                if not nombre_conteo:
                    return JsonResponse({'success': False, 'error': 'Debe ingresar un nombre para el conteo.'})
                
                if not numero_conteo:
                    return JsonResponse({'success': False, 'error': 'Debe seleccionar un número de conteo.'})
                
                try:
                    numero_conteo = int(numero_conteo)
                    if numero_conteo not in [1, 2, 3]:
                        return JsonResponse({'success': False, 'error': 'El número de conteo debe ser 1, 2 o 3.'})
                except (ValueError, TypeError):
                    return JsonResponse({'success': False, 'error': 'Número de conteo inválido.'})
                
                with transaction.atomic():
                    # Verificar si ya existe un conteo con el mismo nombre y número
                    nombre_final = nombre_conteo
                    contador = 1
                    while Conteo.objects.filter(nombre=nombre_final, numero_conteo=numero_conteo).exists():
                        nombre_final = f"{nombre_conteo} ({contador})"
                        contador += 1
                    
                    # Crear el nuevo conteo sin pareja asignada
                    # Guardar los IDs de productos en las observaciones para poder filtrarlos después
                    productos_ids_str = ','.join(str(pid) for pid in producto_ids)
                    conteo = Conteo.objects.create(
                        nombre=nombre_final,
                        numero_conteo=numero_conteo,
                        estado='en_proceso',
                        usuario_creador=request.user,
                        usuario_modificador=request.user,
                        observaciones=f'Conteo creado desde comparativo "{comparativo.nombre}" para recontar productos con diferencias. Productos: {productos_ids_str}'
                    )
                
                mensaje = f'Conteo "{nombre_final}" creado exitosamente con {len(producto_ids)} producto(s).'
                if nombre_final != nombre_conteo:
                    mensaje += f' (El nombre se ajustó a "{nombre_final}" porque ya existía un conteo con el nombre "{nombre_conteo}")'
                mensaje += ' Ahora puede asignar estos productos a parejas en la página de asignación múltiple.'
                
                return JsonResponse({
                    'success': True,
                    'message': mensaje,
                    'conteo_id': conteo.pk
                })
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Error: {str(e)}'})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'})

