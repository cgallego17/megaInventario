from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from django import forms
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .models import Producto
from .forms import ProductoForm, ImportarProductosForm, ImportarProductosAPIForm


@login_required
def lista_productos(request):
    """Lista todos los productos con paginación y filtros"""
    # Mostrar todos los productos por defecto (activos e inactivos)
    productos = Producto.objects.all()
    
    # Filtro opcional para mostrar solo activos
    mostrar_solo_activos = request.GET.get('solo_activos', '').strip()
    if mostrar_solo_activos == 'true' or mostrar_solo_activos == '1':
        productos = productos.filter(activo=True)
    
    # Búsqueda por nombre, código de barras, código, atributo o ID
    busqueda = request.GET.get('busqueda', '').strip()
    if busqueda:
        # Intentar buscar por ID si es un número
        try:
            busqueda_id = int(busqueda)
            productos = productos.filter(
                Q(id=busqueda_id) |
                Q(codigo_barras__icontains=busqueda) |
                Q(codigo_barras__iexact=busqueda) |
                Q(codigo__icontains=busqueda) |
                Q(nombre__icontains=busqueda) |
                Q(marca__icontains=busqueda) |
                Q(descripcion__icontains=busqueda) |
                Q(atributo__icontains=busqueda)
            )
        except ValueError:
            # Si no es un número, buscar por texto (código de barras, código, nombre, descripción, atributo)
            productos = productos.filter(
                Q(codigo_barras__icontains=busqueda) |
                Q(codigo_barras__iexact=busqueda) |
                Q(codigo__icontains=busqueda) |
                Q(nombre__icontains=busqueda) |
                Q(marca__icontains=busqueda) |
                Q(descripcion__icontains=busqueda) |
                Q(atributo__icontains=busqueda)
            )
    
    # Filtros
    marca_filtro = request.GET.get('marca', '').strip()
    if marca_filtro:
        productos = productos.filter(marca__icontains=marca_filtro)
    
    categoria_filtro = request.GET.get('categoria', '').strip()
    if categoria_filtro:
        productos = productos.filter(categoria__icontains=categoria_filtro)
    
    atributo_filtro = request.GET.get('atributo', '').strip()
    if atributo_filtro:
        productos = productos.filter(atributo__icontains=atributo_filtro)
    
    stock_filtro = request.GET.get('stock', '').strip()
    if stock_filtro:
        if stock_filtro == 'con_stock':
            # Productos con stock > 0
            productos_ids = []
            for producto in productos:
                if producto.get_stock_actual() > 0:
                    productos_ids.append(producto.id)
            productos = productos.filter(id__in=productos_ids)
        elif stock_filtro == 'sin_stock':
            # Productos con stock = 0
            productos_ids = []
            for producto in productos:
                if producto.get_stock_actual() == 0:
                    productos_ids.append(producto.id)
            productos = productos.filter(id__in=productos_ids)
    
    precio_min = request.GET.get('precio_min', '').strip()
    if precio_min:
        try:
            precio_min_val = float(precio_min)
            productos = productos.filter(precio__gte=precio_min_val)
        except ValueError:
            pass
    
    precio_max = request.GET.get('precio_max', '').strip()
    if precio_max:
        try:
            precio_max_val = float(precio_max)
            productos = productos.filter(precio__lte=precio_max_val)
        except ValueError:
            pass
    
    # Ordenamiento
    orden = request.GET.get('orden', 'marca')
    orden_opciones = {
        'nombre': 'nombre',
        '-nombre': '-nombre',
        'precio': 'precio',
        '-precio': '-precio',
        'marca': 'marca',
        '-marca': '-marca',
        'categoria': 'categoria',
        '-categoria': '-categoria',
        'codigo_barras': 'codigo_barras',
        '-codigo_barras': '-codigo_barras',
    }
    if orden in orden_opciones:
        productos = productos.order_by(orden_opciones[orden], 'nombre')  # Ordenar por el campo seleccionado y luego por nombre
    else:
        productos = productos.order_by('marca', 'nombre')  # Ordenar por marca y luego por nombre
    
    # Obtener valores únicos para los filtros dropdown (todos los productos)
    marcas = Producto.objects.exclude(marca__isnull=True).exclude(marca='').values_list('marca', flat=True).distinct().order_by('marca')
    categorias = Producto.objects.exclude(categoria__isnull=True).exclude(categoria='').values_list('categoria', flat=True).distinct().order_by('categoria')
    atributos = Producto.objects.exclude(atributo__isnull=True).exclude(atributo='').values_list('atributo', flat=True).distinct().order_by('atributo')
    
    # Paginación
    paginator = Paginator(productos, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Construir URL base para paginación (mantener filtros)
    params = request.GET.copy()
    if 'page' in params:
        del params['page']
    url_params = params.urlencode()
    
    return render(request, 'productos/lista.html', {
        'page_obj': page_obj,
        'busqueda': busqueda,
        'marca_filtro': marca_filtro,
        'categoria_filtro': categoria_filtro,
        'atributo_filtro': atributo_filtro,
        'stock_filtro': stock_filtro,
        'precio_min': precio_min,
        'precio_max': precio_max,
        'orden': orden,
        'marcas': marcas,
        'categorias': categorias,
        'atributos': atributos,
        'url_params': url_params,
    })


@login_required
def crear_producto(request):
    """Crea un nuevo producto"""
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Producto creado exitosamente.')
            return redirect('productos:lista')
    else:
        form = ProductoForm()
    
    return render(request, 'productos/form.html', {'form': form, 'titulo': 'Crear Producto'})


@login_required
def editar_producto(request, pk):
    """Edita un producto existente"""
    producto = get_object_or_404(Producto, pk=pk)
    
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, 'Producto actualizado exitosamente.')
            return redirect('productos:lista')
    else:
        form = ProductoForm(instance=producto)
    
    return render(request, 'productos/form.html', {
        'form': form,
        'producto': producto,
        'titulo': 'Editar Producto'
    })


@login_required
def eliminar_producto(request, pk):
    """Elimina (desactiva) un producto"""
    producto = get_object_or_404(Producto, pk=pk)
    
    if request.method == 'POST':
        producto.activo = False
        producto.save()
        messages.success(request, 'Producto eliminado exitosamente.')
        return redirect('productos:lista')
    
    return render(request, 'productos/eliminar.html', {'producto': producto})


@login_required
def importar_productos(request):
    """Importa productos desde un archivo Excel o CSV con todos los campos necesarios"""
    if request.method == 'POST':
        form = ImportarProductosForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                productos, errores = form.procesar_archivo(request.FILES['archivo'])
                
                # Validar que haya productos para procesar
                if not productos:
                    messages.error(request, "No se encontraron productos válidos para importar. Verifique el archivo.")
                    return render(request, 'productos/importar.html', {'form': form})
                
                # Crear productos
                creados = 0
                actualizados = 0
                errores_guardado = []
                
                with transaction.atomic():
                    for producto_data in productos:
                        try:
                            codigo_barras = producto_data['codigo_barras']
                            
                            # Validar que el código de barras sea único antes de guardar
                            producto_existente = Producto.objects.filter(codigo_barras=codigo_barras).first()
                            
                            if producto_existente:
                                # Actualizar producto existente
                                for key, value in producto_data.items():
                                    setattr(producto_existente, key, value)
                                producto_existente.save()
                                actualizados += 1
                            else:
                                # Crear nuevo producto
                                Producto.objects.create(**producto_data)
                                creados += 1
                                
                        except Exception as e:
                            errores_guardado.append(f"Error al guardar producto {producto_data.get('codigo_barras', 'N/A')}: {str(e)}")
                
                # Construir mensaje de resultado
                mensaje = f"Importación completada: {creados} productos creados, {actualizados} actualizados."
                
                # Agregar información sobre errores
                total_errores = len(errores) + len(errores_guardado)
                if total_errores > 0:
                    mensaje += f" {total_errores} error(es) encontrado(s)."
                    messages.warning(request, mensaje)
                    
                    # Mostrar primeros errores
                    todos_errores = errores + errores_guardado
                    if len(todos_errores) <= 10:
                        for error in todos_errores:
                            messages.info(request, error)
                    else:
                        for error in todos_errores[:10]:
                            messages.info(request, error)
                        messages.info(request, f"... y {len(todos_errores) - 10} error(es) más")
                else:
                    messages.success(request, mensaje)
                
                return redirect('productos:lista')
                
            except forms.ValidationError as e:
                messages.error(request, f"Error de validación: {str(e)}")
            except Exception as e:
                messages.error(request, f"Error al importar: {str(e)}")
                import traceback
                if request.user.is_superuser:  # Solo mostrar traceback a superusuarios
                    messages.error(request, f"Detalles: {traceback.format_exc()}")
    else:
        form = ImportarProductosForm()
    
    return render(request, 'productos/importar.html', {'form': form})


@login_required
def importar_productos_api(request):
    """Importa productos desde una API externa"""
    if request.method == 'POST':
        form = ImportarProductosAPIForm(request.POST)
        if form.is_valid():
            try:
                url_api = form.cleaned_data['url_api']
                headers = form.cleaned_data.get('headers_json')
                metodo = form.cleaned_data.get('metodo', 'GET')
                mapeo_personalizado = form.cleaned_data.get('mapeo_personalizado_json')
                
                # Importar función desde el script
                import sys
                import os
                script_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'importar_productos_api.py')
                if os.path.exists(script_path):
                    # Ejecutar importación
                    from importar_productos_api import importar_desde_api
                    
                    creados, actualizados, errores = importar_desde_api(
                        url_api=url_api,
                        headers=headers,
                        metodo=metodo,
                        mapeo_personalizado=mapeo_personalizado
                    )
                    
                    mensaje = f"Importación desde API completada: {creados} productos creados, {actualizados} actualizados."
                    if errores:
                        mensaje += f" {len(errores)} error(es) encontrado(s)."
                        messages.warning(request, mensaje)
                        if len(errores) <= 10:
                            for error in errores:
                                messages.info(request, error)
                        else:
                            for error in errores[:10]:
                                messages.info(request, error)
                            messages.info(request, f"... y {len(errores) - 10} error(es) más")
                    else:
                        messages.success(request, mensaje)
                    
                    return redirect('productos:lista')
                else:
                    messages.error(request, "Error: No se encontró el módulo de importación desde API")
            except Exception as e:
                messages.error(request, f"Error al importar desde API: {str(e)}")
                import traceback
                if request.user.is_superuser:
                    messages.error(request, f"Detalles: {traceback.format_exc()}")
    else:
        form = ImportarProductosAPIForm()
    
    return render(request, 'productos/importar_api.html', {'form': form})


@login_required
def detalle_producto(request, pk):
    """Muestra el detalle de un producto"""
    producto = get_object_or_404(Producto, pk=pk)
    return render(request, 'productos/detalle.html', {'producto': producto})


@login_required
def descargar_plantilla_importacion(request):
    """Descarga una plantilla Excel para importar productos con todos los campos necesarios"""
    # Crear datos para la plantilla
    datos_plantilla = [
        {
            'codigo_barras': '1234567890123',
            'codigo': 'P001',
            'nombre': 'Producto Ejemplo 1',
            'marca': 'Marca A',
            'descripcion': 'Descripción del producto ejemplo 1',
            'categoria': 'Categoría 1',
            'atributo': 'Atributo ejemplo',
            'precio': 100.00,
            'unidad_medida': 'UN',
            'activo': 'Sí',
        },
        {
            'codigo_barras': '9876543210987',
            'codigo': 'P002',
            'nombre': 'Producto Ejemplo 2',
            'marca': 'Marca B',
            'descripcion': 'Descripción del producto ejemplo 2',
            'categoria': 'Categoría 2',
            'atributo': '',
            'precio': 250.50,
            'unidad_medida': 'KG',
            'activo': 'Sí',
        },
        {
            'codigo_barras': '5555555555555',
            'codigo': 'P003',
            'nombre': 'Producto Ejemplo 3',
            'marca': '',
            'descripcion': '',
            'categoria': 'Categoría 1',
            'atributo': '',
            'precio': 0.00,
            'unidad_medida': 'UN',
            'activo': 'Sí',
        },
    ]
    
    df = pd.DataFrame(datos_plantilla)
    
    # Crear archivo Excel en memoria
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Productos', index=False)
        
        # Obtener la hoja para formatear
        worksheet = writer.sheets['Productos']
        
        # Importar estilos de openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 20,  # codigo_barras
            'B': 15,  # codigo
            'C': 40,  # nombre
            'D': 15,  # marca
            'E': 50,  # descripcion
            'F': 20,  # categoria
            'G': 20,  # atributo
            'H': 15,  # precio
            'I': 15,  # unidad_medida
            'J': 10,  # activo
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Formatear encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Formatear filas de datos
        for row_idx in range(2, len(df) + 2):
            for col_letter in column_widths.keys():
                cell = worksheet[f'{col_letter}{row_idx}']
                cell.border = border
                
                # Formatear según el tipo de columna
                if col_letter == 'H':  # Precio
                    if cell.value and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                else:  # Columnas de texto
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Agregar hoja de instrucciones
        worksheet_instrucciones = writer.book.create_sheet("Instrucciones", 0)
        
        # Contenido de instrucciones
        instrucciones = [
            ['INSTRUCCIONES PARA IMPORTAR PRODUCTOS', ''],
            ['', ''],
            ['COLUMNAS REQUERIDAS:', ''],
            ['codigo_barras', 'Código de barras único del producto (REQUERIDO)'],
            ['nombre', 'Nombre del producto (REQUERIDO)'],
            ['', ''],
            ['COLUMNAS OPCIONALES:', ''],
            ['codigo', 'Código interno del producto'],
            ['marca', 'Marca del producto'],
            ['descripcion', 'Descripción detallada del producto'],
            ['categoria', 'Categoría del producto'],
            ['atributo', 'Atributo o característica adicional'],
            ['precio', 'Precio unitario (default: 0.00)'],
            ['unidad_medida', 'Unidad de medida (default: UN)'],
            ['activo', 'Estado activo: Sí/No (default: Sí)'],
            ['', ''],
            ['NOTAS IMPORTANTES:', ''],
            ['', '• El código de barras debe ser único'],
            ['', '• El nombre es obligatorio'],
            ['', '• Los valores vacíos se manejarán automáticamente'],
            ['', '• El precio debe ser un número (puede ser 0)'],
            ['', '• Si un producto ya existe (mismo código de barras), se actualizará'],
            ['', '• Formatos soportados: Excel (.xlsx, .xls) o CSV'],
            ['', '• Los nombres de columnas son flexibles (acepta variantes)'],
        ]
        
        # Escribir instrucciones
        for row_idx, (col1, col2) in enumerate(instrucciones, start=1):
            cell1 = worksheet_instrucciones.cell(row=row_idx, column=1, value=col1)
            cell2 = worksheet_instrucciones.cell(row=row_idx, column=2, value=col2)
            
            # Formatear título
            if row_idx == 1:
                cell1.font = Font(bold=True, size=14, color="366092")
                cell1.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
                worksheet_instrucciones.merge_cells(f'A{row_idx}:B{row_idx}')
                cell1.alignment = Alignment(horizontal="center", vertical="center")
            # Formatear subtítulos
            elif col1 in ['COLUMNAS REQUERIDAS:', 'COLUMNAS OPCIONALES:', 'NOTAS IMPORTANTES:']:
                cell1.font = Font(bold=True, size=11, color="000000")
                cell1.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell2.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            # Formatear nombres de columnas
            elif col1 and col1 not in ['', 'COLUMNAS REQUERIDAS:', 'COLUMNAS OPCIONALES:', 'NOTAS IMPORTANTES:'] and not col1.startswith('•'):
                cell1.font = Font(bold=True, color="366092")
                cell1.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            cell1.border = border
            cell2.border = border
            cell1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Ajustar anchos de columna en hoja de instrucciones
        worksheet_instrucciones.column_dimensions['A'].width = 25
        worksheet_instrucciones.column_dimensions['B'].width = 60
        
        # Agregar copyright en la hoja de Productos
        copyright_row = len(df) + 2
        from datetime import datetime
        current_year = datetime.now().year
        worksheet.merge_cells(f'A{copyright_row}:J{copyright_row}')
        copyright_cell = worksheet[f'A{copyright_row}']
        copyright_cell.value = f'© {current_year} Todos los derechos reservados por megadominio.co'
        copyright_cell.alignment = Alignment(horizontal="center", vertical="center")
        copyright_font = Font(size=9, italic=True, color="808080")
        copyright_cell.font = copyright_font
        
        # Agregar copyright en la hoja de Instrucciones
        instrucciones_copyright_row = len(instrucciones) + 2
        worksheet_instrucciones.merge_cells(f'A{instrucciones_copyright_row}:B{instrucciones_copyright_row}')
        instrucciones_copyright_cell = worksheet_instrucciones[f'A{instrucciones_copyright_row}']
        instrucciones_copyright_cell.value = f'© {current_year} Todos los derechos reservados por megadominio.co'
        instrucciones_copyright_cell.alignment = Alignment(horizontal="center", vertical="center")
        instrucciones_copyright_cell.font = copyright_font
    
    output.seek(0)
    
    # Crear la respuesta HTTP
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="plantilla_importacion_productos_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    return response


@login_required
def exportar_productos(request):
    """Exporta todos los productos (activos e inactivos) a Excel con todos los campos necesarios"""
    # Obtener todos los productos (activos e inactivos)
    productos = Producto.objects.all().order_by('nombre')
    
    # Preparar datos para el DataFrame
    # Usar nombres de columnas compatibles con la importación (minúsculas, sin espacios o con guiones bajos)
    datos = []
    for producto in productos:
        datos.append({
            'codigo_barras': producto.codigo_barras,
            'codigo': producto.codigo or '',
            'nombre': producto.nombre,
            'marca': producto.marca or '',
            'descripcion': producto.descripcion or '',
            'categoria': producto.categoria or '',
            'atributo': producto.atributo or '',
            'precio': float(producto.precio),
            'unidad_medida': producto.unidad_medida,
            'stock_actual': producto.get_stock_actual(),  # Stock calculado dinámicamente
            'activo': 'Sí' if producto.activo else 'No',
            'fecha_creacion': producto.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if producto.fecha_creacion else '',
            'fecha_actualizacion': producto.fecha_actualizacion.strftime('%Y-%m-%d %H:%M:%S') if producto.fecha_actualizacion else '',
        })
    
    # Crear DataFrame
    df = pd.DataFrame(datos)
    
    # Crear archivo Excel en memoria
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Productos', index=False)
        
        # Obtener la hoja para formatear
        worksheet = writer.sheets['Productos']
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 20,  # codigo_barras
            'B': 15,  # codigo
            'C': 40,  # nombre
            'D': 15,  # marca
            'E': 50,  # descripcion
            'F': 20,  # categoria
            'G': 20,  # atributo
            'H': 15,  # precio
            'I': 15,  # unidad_medida
            'J': 15,  # stock_actual
            'K': 10,  # activo
            'L': 20,  # fecha_creacion
            'M': 20,  # fecha_actualizacion
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Formatear encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Formatear filas de datos
        for row_idx in range(2, len(df) + 2):
            for col_letter in column_widths.keys():
                cell = worksheet[f'{col_letter}{row_idx}']
                cell.border = border
                
                # Formatear según el tipo de columna
                if col_letter == 'H':  # Precio
                    if cell.value and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_letter == 'J':  # Stock Actual
                    if cell.value and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_letter in ['L', 'M']:  # Fechas
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:  # Columnas de texto
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Agregar fila de totales
        if len(datos) > 0:
            total_row = len(df) + 2
            total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            total_font = Font(bold=True, size=11)
            
            # Total productos
            worksheet.merge_cells(f'A{total_row}:C{total_row}')
            total_cell = worksheet[f'A{total_row}']
            total_cell.value = f'TOTAL PRODUCTOS: {len(datos)}'
            total_cell.fill = total_fill
            total_cell.font = total_font
            total_cell.alignment = Alignment(horizontal="right", vertical="center")
            total_cell.border = border
            
            # Total stock
            total_stock = sum(p.get_stock_actual() for p in productos)
            stock_cell = worksheet[f'J{total_row}']
            stock_cell.value = total_stock
            stock_cell.fill = total_fill
            stock_cell.font = total_font
            stock_cell.number_format = '#,##0'
            stock_cell.alignment = Alignment(horizontal="right", vertical="center")
            stock_cell.border = border
            
            # Total valor (precio * stock) - en columna I (unidad_medida) o crear nueva columna
            total_valor = sum(float(p.precio) * p.get_stock_actual() for p in productos)
            worksheet.merge_cells(f'I{total_row}:K{total_row}')
            valor_cell = worksheet[f'I{total_row}']
            valor_cell.value = f'Valor Total Inventario: ${total_valor:,.2f}'
            valor_cell.fill = total_fill
            valor_cell.font = total_font
            valor_cell.alignment = Alignment(horizontal="right", vertical="center")
            valor_cell.border = border
        
        # Agregar copyright al final
        copyright_row = len(df) + 3
        from datetime import datetime
        current_year = datetime.now().year
        worksheet.merge_cells(f'A{copyright_row}:M{copyright_row}')
        copyright_cell = worksheet[f'A{copyright_row}']
        copyright_cell.value = f'© {current_year} Todos los derechos reservados por megadominio.co'
        copyright_cell.alignment = Alignment(horizontal="center", vertical="center")
        copyright_font = Font(size=9, italic=True, color="808080")
        copyright_cell.font = copyright_font
    
    output.seek(0)
    
    # Crear la respuesta HTTP
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="productos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response

